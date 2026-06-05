# ================================================================
#  Low Consumption Consumer Tracker — Desktop App
#  Office-wise split into 10 separate Excel files
#  Requires Python 3.8+  |  run: python merge_app.py
# ================================================================

import subprocess, sys, os

def _pip(*pkgs):
    subprocess.check_call([sys.executable, "-m", "pip", "install", *pkgs, "-q"])

try:
    import pandas as pd
except ImportError:
    print("Installing pandas..."); _pip("pandas"); import pandas as pd
try:
    import openpyxl
except ImportError:
    print("Installing openpyxl..."); _pip("openpyxl"); import openpyxl

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading, re, traceback
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ================================================================
#  OFFICE CODES — first 6 digits of হিসাব নম্বর
# ================================================================
OFFICES = {
    "101901": "HQ",
    "101902": "Daulotpur",
    "101903": "Kumarkhali",
    "101904": "Mirpur",
    "101905": "Bheramara",
    "101906": "Swastipur",
    "101907": "Khoksha",
    "101908": "Poradah",
    "101909": "Pragpur",
    "101910": "Panti",
}

OUTPUT_SUBFOLDER = "Low usage Industry Consumer Officewise"

# ================================================================
#  PROCESSING ENGINE
# ================================================================
MONTH_BN = {
    1:"জানুয়ারি", 2:"ফেব্রুয়ারি", 3:"মার্চ",
    4:"এপ্রিল",   5:"মে",          6:"জুন",
    7:"জুলাই",    8:"আগস্ট",       9:"সেপ্টেম্বর",
    10:"অক্টোবর", 11:"নভেম্বর",   12:"ডিসেম্বর",
}
BENGALI_DIGITS = str.maketrans("0123456789", "০১২৩৪৫৬৭৮৯")
def to_bn_year(y): return str(y).translate(BENGALI_DIGITS)

def month_label_from_series(series):
    col = pd.to_datetime(series.dropna(), errors="coerce").dropna()
    if col.empty: return None
    p = col.dt.to_period("M").mode()[0]
    return f"{MONTH_BN[p.month]} {to_bn_year(p.year)}"

def month_label_from_filename(name):
    en = {"january":"জানুয়ারি","february":"ফেব্রুয়ারি","march":"মার্চ",
          "april":"এপ্রিল","may":"মে","june":"জুন","july":"জুলাই",
          "august":"আগস্ট","september":"সেপ্টেম্বর","october":"অক্টোবর",
          "november":"নভেম্বর","december":"ডিসেম্বর"}
    m = re.search(
        r"(january|february|march|april|may|june|july|august|"
        r"september|october|november|december)[.\s_-]*(\d{2,4})",
        name, re.IGNORECASE)
    if not m: return None
    yr = m.group(2); yr = "20"+yr if len(yr)==2 else yr
    return f"{en[m.group(1).lower()]} {to_bn_year(int(yr))}"

ALIASES = {
    "হিসাব নম্বর":            ["হিসাব নম্বর","হিসাব_নম্বর"],
    "গ্রাহকের নাম":           ["গ্রাহকের নাম","গ্রাহকের_নাম"],
    "ট্যারিফ":                ["ট্যারিফ","tariff"],
    "চুক্তিবদ্ধ লোড (কি.ও.)": ["চুক্তিবদ্ধ লোড (কি.ও.)","চুক্তিবদ্ধ লোড"],
    "মোট ব্যবহৃত (কি.ও.ঘ.)":  ["মোট ব্যবহৃত (কি.ও.ঘ.)","মোট ব্যবহৃত"],
    "বিল মাস":                ["বিল মাস"],
    "মোবাইল নম্বর":           ["মোবাইল নম্বর","মোবাইল_নম্বর","mobile"],
}

def find_col(columns, key):
    for alias in ALIASES.get(key, [key]):
        for c in columns:
            if str(c).strip() == alias.strip(): return c
    return None

def clean_mobile(val):
    if pd.isna(val) or str(val).strip() in ("","nan","NaN"): return "N/A"
    try:
        s = str(int(float(str(val))))
        return ("0"+s) if len(s)==10 else s
    except: return str(val).strip()

def month_sort_key(label):
    rev = {v:k for k,v in MONTH_BN.items()}
    parts = label.split()
    if len(parts)>=2:
        mn = rev.get(parts[0], 0)
        try: yr = int(parts[-1].translate(str.maketrans("০১২৩৪৫৬৭৮৯","0123456789")))
        except: yr = 0
        return (yr, mn)
    return (9999,99)


def process_files(file_paths, log):
    """Read all monthly Excel files and merge into one master DataFrame.
       Missing monthly values are filled with 0 (not 'N/A'), so no blank cells."""
    month_dfs, month_order = {}, []

    for path in file_paths:
        fname = Path(path).name
        log(f"🔍 Processing: {fname}")
        try:
            xl  = pd.ExcelFile(path)
            df  = pd.read_excel(xl, sheet_name=xl.sheet_names[0], header=0)
            df.columns = [str(c).strip() for c in df.columns]

            c_hesab  = find_col(df.columns, "হিসাব নম্বর")
            c_name   = find_col(df.columns, "গ্রাহকের নাম")
            c_tariff = find_col(df.columns, "ট্যারিফ")
            c_load   = find_col(df.columns, "চুক্তিবদ্ধ লোড (কি.ও.)")
            c_usage  = find_col(df.columns, "মোট ব্যবহৃত (কি.ও.ঘ.)")
            c_bilmas = find_col(df.columns, "বিল মাস")
            c_mobile = find_col(df.columns, "মোবাইল নম্বর")

            missing = [k for k,v in {"হিসাব নম্বর":c_hesab,
                                      "গ্রাহকের নাম":c_name,
                                      "মোট ব্যবহৃত":c_usage}.items() if v is None]
            if missing:
                log(f"  ⚠️  Skipped — missing columns: {missing}"); continue

            label = (month_label_from_series(df[c_bilmas]) if c_bilmas else None) \
                    or month_label_from_filename(fname) \
                    or fname.replace(".xlsx","")
            log(f"  📅 Month: {label}")

            sub = df[[c_hesab, c_name]].copy()
            sub.columns = ["হিসাব নম্বর","গ্রাহকের নাম"]
            if c_tariff:  sub["ট্যারিফ"]                = df[c_tariff].values
            if c_load:    sub["চুক্তিবদ্ধ লোড (কি.ও.)"] = df[c_load].values
            if c_mobile:  sub["মোবাইল নম্বর"]            = df[c_mobile].values
            sub[f"usage_{label}"] = df[c_usage].values
            sub["হিসাব নম্বর"] = sub["হিসাব নম্বর"].astype(str).str.strip()

            if label in month_dfs:
                month_dfs[label] = pd.concat([month_dfs[label], sub], ignore_index=True)
            else:
                month_dfs[label] = sub
                month_order.append(label)

            log(f"  ✅ {len(sub)} rows loaded.")
        except Exception as e:
            log(f"  ❌ Error: {e}")

    if not month_dfs:
        raise RuntimeError("No valid data found in any file.")

    month_order.sort(key=month_sort_key)
    log(f"\n📊 Months: {', '.join(month_order)}")

    all_rec = pd.concat(list(month_dfs.values()), ignore_index=True)
    master  = all_rec.drop_duplicates(subset="হিসাব নম্বর")[
        ["হিসাব নম্বর","গ্রাহকের নাম"]].copy()

    for col in ["ট্যারিফ","চুক্তিবদ্ধ লোড (কি.ও.)","মোবাইল নম্বর"]:
        if col in all_rec.columns:
            lat = (all_rec[["হিসাব নম্বর",col]]
                   .dropna(subset=[col])
                   .drop_duplicates("হিসাব নম্বর", keep="last"))
            master = master.merge(lat, on="হিসাব নম্বর", how="left")

    usage_cols = []
    for m in month_order:
        dm = (month_dfs[m][["হিসাব নম্বর",f"usage_{m}"]]
              .dropna(subset=["হিসাব নম্বর"])
              .drop_duplicates("হিসাব নম্বর"))
        master = master.merge(dm, on="হিসাব নম্বর", how="left")
        usage_cols.append(f"usage_{m}")

    # Fill missing monthly usage with 0 and convert to numeric so cells never appear blank
    for uc in usage_cols:
        master[uc] = pd.to_numeric(master[uc], errors="coerce").fillna(0)

    if "মোবাইল নম্বর" in master.columns:
        master["মোবাইল নম্বর"] = master["মোবাইল নম্বর"].apply(clean_mobile)

    master = master.sort_values("গ্রাহকের নাম").reset_index(drop=True)
    has_mob = "মোবাইল নম্বর" in master.columns
    log(f"\n✅ Unique customers across all months: {len(master)}")

    return master, len(master), month_order, usage_cols, has_mob


def build_formatted_workbook(master, month_order, usage_cols, has_mob, sheet_title):
    """Build a styled openpyxl Workbook. Returns (wb, month_col_indices)."""
    wb = Workbook(); ws = wb.active; ws.title = sheet_title[:31]  # Excel limit

    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", start_color="1F4E79")
    sub_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    sub_fill  = PatternFill("solid", start_color="2E75B6")
    alt_fill  = PatternFill("solid", start_color="DCE6F1")
    wht_fill  = PatternFill("solid", start_color="FFFFFF")
    dat_font  = Font(name="Arial", size=10)
    ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
    la = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    bdr = Border(left=Side(style="thin"), right=Side(style="thin"),
                 top=Side(style="thin"),  bottom=Side(style="thin"))

    static = ["হিসাব নম্বর","গ্রাহকের নাম"]
    for c in ["ট্যারিফ","চুক্তিবদ্ধ লোড (কি.ও.)"]:
        if c in master.columns: static.append(c)
    ns, nm  = len(static), len(month_order)
    mob_col = ns + nm + 1
    month_col_indices = [ns + 1 + j for j in range(nm)]

    def hcell(r,c,v,font,fill):
        cell = ws.cell(r,c,v)
        cell.font,cell.fill,cell.alignment,cell.border = font,fill,ca,bdr
        return cell

    for i,col in enumerate(static,1):
        hcell(1,i,col,hdr_font,hdr_fill)
        ws.merge_cells(start_row=1,start_column=i,end_row=2,end_column=i)

    if nm:
        hcell(1,ns+1,"মোট ব্যবহৃত (কি.ও.ঘ.) — মাসওয়ারী",hdr_font,hdr_fill)
        if nm>1: ws.merge_cells(start_row=1,start_column=ns+1,end_row=1,end_column=ns+nm)
        for j,m in enumerate(month_order):
            hcell(2,ns+1+j,m,sub_font,sub_fill)

    if has_mob:
        hcell(1,mob_col,"মোবাইল নম্বর",hdr_font,hdr_fill)
        ws.merge_cells(start_row=1,start_column=mob_col,end_row=2,end_column=mob_col)
        ws.cell(2,mob_col).border = bdr

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25

    ordered = static + usage_cols + (["মোবাইল নম্বর"] if has_mob else [])
    for ri,(_,row) in enumerate(master.iterrows(),3):
        fill = alt_fill if ri%2==0 else wht_fill
        for ci,col in enumerate(ordered,1):
            v = row.get(col, 0 if col in usage_cols else "N/A")
            if pd.isna(v): v = 0 if col in usage_cols else "N/A"
            # ensure numeric usage values are stored as numbers (no blank cells)
            if col in usage_cols:
                try: v = float(v)
                except (TypeError, ValueError): v = 0
                if v == int(v): v = int(v)
            cell = ws.cell(ri,ci,v)
            cell.font,cell.fill,cell.border = dat_font,fill,bdr
            cell.alignment = la if col=="গ্রাহকের নাম" else ca

    widths = {"হিসাব নম্বর":20,"গ্রাহকের নাম":34,
              "ট্যারিফ":10,"চুক্তিবদ্ধ লোড (কি.ও.)":22,"মোবাইল নম্বর":16}
    for i,col in enumerate(ordered,1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col,17)

    ws.freeze_panes = "C3"
    for j,m in enumerate(month_order):
        ws.cell(2, ns+1+j).value = m

    return wb, month_col_indices


# ================================================================
#  LOW-USAGE FILTER
# ================================================================
HEADER_ROWS = 2

def row_qualifies(ws_row, month_col_indices, threshold):
    for col_idx in month_col_indices:
        val = ws_row[col_idx - 1].value
        if val is not None:
            try:
                if float(val) <= threshold:
                    return True
            except (TypeError, ValueError):
                pass
    return False


def filter_low_usage(wb, month_col_indices, threshold, log, label=""):
    """Delete non-qualifying rows in-place. Returns (total, kept, removed)."""
    ws = wb.active
    rows_to_delete = []
    total_data_rows = 0
    for row_num in range(HEADER_ROWS + 1, ws.max_row + 1):
        ws_row = ws[row_num]
        if all(cell.value is None for cell in ws_row):
            continue
        total_data_rows += 1
        if not row_qualifies(ws_row, month_col_indices, threshold):
            rows_to_delete.append(row_num)

    kept = total_data_rows - len(rows_to_delete)
    removed = len(rows_to_delete)
    for row_num in reversed(rows_to_delete):
        ws.delete_rows(row_num)

    tag = f" [{label}]" if label else ""
    log(f"🔻 Filter{tag} (≤ {threshold}): total={total_data_rows}, kept={kept}, removed={removed}")
    return total_data_rows, kept, removed


def subset_by_office(master, office_code):
    hesab = master["হিসাব নম্বর"].astype(str).str.strip()
    return master[hesab.str[:6] == office_code].copy()


# ================================================================
#  GUI
# ================================================================
DARK_BLUE  = "#1F4E79"; MED_BLUE = "#2E75B6"; LIGHT_BLUE = "#D6E4F0"
WHITE      = "#FFFFFF"; ACCENT   = "#E8F4FD"; TEXT_DARK  = "#1A1A2E"
SUCCESS    = "#1A7A4A"; ERROR    = "#C0392B"

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Low Consumption Consumer Tracker")
        self.geometry("820x720")
        self.minsize(720, 620)
        self.configure(bg=WHITE)
        self.resizable(True, True)

        self.files        = []
        self.output_dir   = tk.StringVar(value=str(Path.home() / "Desktop"))
        self.apply_filter = tk.BooleanVar(value=True)
        self.threshold    = tk.StringVar(value="150")
        self.processing   = False

        self._build_ui()
        self._center()

    def _build_ui(self):
        hdr = tk.Frame(self, bg=DARK_BLUE, height=70)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Label(hdr, text="⚡  Low Consumption Consumer Tracker",
                 font=("Arial", 18, "bold"), bg=DARK_BLUE, fg=WHITE
                 ).pack(side="left", padx=20, pady=15)
        tk.Label(hdr, text="Office Wise",
                 font=("Arial", 10), bg=DARK_BLUE, fg=LIGHT_BLUE
                 ).pack(side="right", padx=20)

        body = tk.Frame(self, bg=WHITE)
        body.pack(fill="both", expand=True, padx=20, pady=16)

        self._section(body, "① Excel ফাইল নির্বাচন করুন")
        file_frame = tk.Frame(body, bg=WHITE)
        file_frame.pack(fill="both", expand=True, pady=(0,10))
        list_wrap = tk.Frame(file_frame, bg=LIGHT_BLUE, bd=1, relief="solid")
        list_wrap.pack(fill="both", expand=True, side="left")
        self.listbox = tk.Listbox(
            list_wrap, bg=ACCENT, fg=TEXT_DARK,
            font=("Consolas", 9), selectbackground=MED_BLUE,
            selectforeground=WHITE, bd=0, highlightthickness=0,
            activestyle="none", height=8)
        sb = ttk.Scrollbar(list_wrap, orient="vertical", command=self.listbox.yview)
        self.listbox.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self.listbox.pack(fill="both", expand=True, padx=4, pady=4)
        self._show_placeholder()

        btn_col = tk.Frame(file_frame, bg=WHITE, width=120)
        btn_col.pack(side="right", fill="y", padx=(10,0))
        btn_col.pack_propagate(False)
        self._btn(btn_col, "➕  Add Files",  self._add_files,   MED_BLUE).pack(fill="x", pady=4)
        self._btn(btn_col, "🗑  Remove",     self._remove_file, "#6C757D").pack(fill="x", pady=4)
        self._btn(btn_col, "✖  Clear All",  self._clear_files, "#6C757D").pack(fill="x", pady=4)
        tk.Frame(btn_col, bg=WHITE, height=10).pack()
        self.count_lbl = tk.Label(btn_col, text="0 files", bg=WHITE,
                                  fg=MED_BLUE, font=("Arial", 9, "bold"))
        self.count_lbl.pack()

        self._section(body, "② আউটপুট ফোল্ডার")
        out_frame = tk.Frame(body, bg=WHITE)
        out_frame.pack(fill="x", pady=(0,10))
        tk.Label(out_frame, text="Folder:", bg=WHITE, fg=TEXT_DARK,
                 font=("Arial", 9)).grid(row=0, column=0, sticky="w", padx=(0,6))
        tk.Entry(out_frame, textvariable=self.output_dir,
                 font=("Arial", 9), width=48, fg=TEXT_DARK,
                 relief="solid", bd=1).grid(row=0, column=1, sticky="ew", padx=(0,6))
        self._btn(out_frame, "📁 Browse", self._browse_dir, MED_BLUE,
                  padx=8, pady=3).grid(row=0, column=2)
        tk.Label(out_frame,
                 text=f"ফাইলগুলো '{OUTPUT_SUBFOLDER}' সাব-ফোল্ডারে সংরক্ষিত হবে",
                 bg=WHITE, fg="#6C757D", font=("Arial", 8, "italic")
                 ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(4,0))
        out_frame.columnconfigure(1, weight=1)

        self._section(body, "③ লো-ইউসেজ ফিল্টার")
        flt_frame = tk.Frame(body, bg=WHITE)
        flt_frame.pack(fill="x", pady=(0,10))
        self.flt_check = tk.Checkbutton(
            flt_frame,
            text="যেকোনো মাসে ব্যবহার ≤ থ্রেশহোল্ড — শুধু এমন গ্রাহকদের রাখুন",
            variable=self.apply_filter,
            bg=WHITE, fg=TEXT_DARK, font=("Arial", 9),
            activebackground=WHITE, anchor="w",
            command=self._toggle_threshold_state)
        self.flt_check.grid(row=0, column=0, columnspan=3, sticky="w")
        tk.Label(flt_frame, text="Threshold (কি.ও.ঘ.):", bg=WHITE, fg=TEXT_DARK,
                 font=("Arial", 9)).grid(row=1, column=0, sticky="w",
                                         padx=(20,6), pady=(4,0))
        self.threshold_entry = tk.Entry(
            flt_frame, textvariable=self.threshold,
            font=("Arial", 9), width=10, fg=TEXT_DARK,
            relief="solid", bd=1, justify="center")
        self.threshold_entry.grid(row=1, column=1, sticky="w", pady=(4,0))
        tk.Label(flt_frame, text="(ডিফল্ট 150)",
                 bg=WHITE, fg="#6C757D", font=("Arial", 8, "italic")
                 ).grid(row=1, column=2, sticky="w", padx=(8,0), pady=(4,0))

        office_summary = "অফিসসমূহ: " + ", ".join(
            f"{name} ({code})" for code, name in OFFICES.items())
        tk.Label(body, text=office_summary, bg=WHITE, fg="#6C757D",
                 font=("Arial", 8), wraplength=760, justify="left"
                 ).pack(fill="x", pady=(4,0))

        run_wrap = tk.Frame(body, bg=WHITE)
        run_wrap.pack(fill="x", pady=8)
        self.run_btn = self._btn(
            run_wrap, "▶  Run", self._run,
            DARK_BLUE, font=("Arial", 12, "bold"), padx=30, pady=10)
        self.run_btn.pack(side="left")
        self.progress = ttk.Progressbar(run_wrap, mode="indeterminate", length=200)
        self.progress.pack(side="left", padx=16, pady=4)
        self.status_lbl = tk.Label(run_wrap, text="", bg=WHITE,
                                   font=("Arial", 9, "italic"), fg=MED_BLUE)
        self.status_lbl.pack(side="left")

        self._section(body, "④ লগ")
        log_wrap = tk.Frame(body, bg=LIGHT_BLUE, bd=1, relief="solid")
        log_wrap.pack(fill="both", expand=True)
        self.log_text = tk.Text(log_wrap, bg="#0D1117", fg="#C9D1D9",
                                font=("Consolas", 9), bd=0, highlightthickness=0,
                                state="disabled", height=8)
        log_sb = ttk.Scrollbar(log_wrap, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_sb.set)
        log_sb.pack(side="right", fill="y")
        self.log_text.pack(fill="both", expand=True, padx=2, pady=2)
        self.log_text.tag_config("ok",   foreground="#3FB950")
        self.log_text.tag_config("err",  foreground="#F85149")
        self.log_text.tag_config("info", foreground="#79C0FF")
        self.log_text.tag_config("warn", foreground="#E3B341")

    def _section(self, parent, text):
        f = tk.Frame(parent, bg=WHITE); f.pack(fill="x", pady=(6,4))
        tk.Label(f, text=text, bg=WHITE, fg=DARK_BLUE,
                 font=("Arial", 10, "bold")).pack(side="left")
        tk.Frame(f, bg=LIGHT_BLUE, height=2).pack(side="left", fill="x",
                                                   expand=True, padx=8, pady=8)

    def _btn(self, parent, text, cmd, bg, font=("Arial",9,"bold"),
             padx=10, pady=6, **kw):
        return tk.Button(parent, text=text, command=cmd, bg=bg, fg=WHITE,
                         font=font, relief="flat", cursor="hand2",
                         activebackground=DARK_BLUE, activeforeground=WHITE,
                         padx=padx, pady=pady, **kw)

    def _center(self):
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        w, h   = self.winfo_width(), self.winfo_height()
        self.geometry(f"+{(sw-w)//2}+{(sh-h)//2}")

    def _toggle_threshold_state(self):
        state = "normal" if self.apply_filter.get() else "disabled"
        try: self.threshold_entry.configure(state=state)
        except Exception: pass

    def _show_placeholder(self):
        if not self.files:
            self.listbox.configure(state="normal")
            self.listbox.delete(0, "end")
            self.listbox.insert("end", "  (এখানে কোনো ফাইল নেই — 'Add Files' দিয়ে যোগ করুন)")
            self.listbox.configure(fg="#9CA3AF")

    def _refresh_list(self):
        self.listbox.configure(state="normal", fg=TEXT_DARK)
        self.listbox.delete(0, "end")
        if not self.files: self._show_placeholder()
        else:
            for i, p in enumerate(self.files, 1):
                self.listbox.insert("end", f"  {i:>2}.  {Path(p).name}")
        self.count_lbl.config(text=f"{len(self.files)} file(s)")

    def _add_files(self):
        paths = filedialog.askopenfilenames(
            title="Excel ফাইল নির্বাচন করুন",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files","*.*")])
        for p in paths:
            if p not in self.files: self.files.append(p)
        self._refresh_list()

    def _remove_file(self):
        sel = self.listbox.curselection()
        if not sel: return
        self.files.pop(sel[0])
        self._refresh_list()

    def _clear_files(self): self.files.clear(); self._refresh_list()

    def _browse_dir(self):
        d = filedialog.askdirectory(title="আউটপুট ফোল্ডার বেছে নিন")
        if d: self.output_dir.set(d)

    def _log(self, msg, tag=""):
        def _write():
            self.log_text.configure(state="normal")
            if not tag:
                if "✅" in msg or "🎉" in msg: t = "ok"
                elif "❌" in msg:              t = "err"
                elif "⚠️" in msg:             t = "warn"
                elif "🔍" in msg or "📅" in msg or "📊" in msg \
                  or "📥" in msg or "🔻" in msg or "🏢" in msg: t = "info"
                else: t = ""
            else: t = tag
            self.log_text.insert("end", msg+"\n", t)
            self.log_text.see("end")
            self.log_text.configure(state="disabled")
        self.after(0, _write)

    def _set_status(self, text, color=MED_BLUE):
        self.after(0, lambda: self.status_lbl.config(text=text, fg=color))

    def _run(self):
        if self.processing: return
        if not self.files:
            messagebox.showwarning("ফাইল নেই", "অনুগ্রহ করে অন্তত একটি Excel ফাইল যোগ করুন।")
            return
        out_folder = self.output_dir.get().strip()
        if not os.path.isdir(out_folder):
            messagebox.showerror("ফোল্ডার পাওয়া যায়নি",
                                 f"এই ফোল্ডারটি বিদ্যমান নেই:\n{out_folder}")
            return
        apply_flt = self.apply_filter.get()
        thr_val   = 150.0
        if apply_flt:
            try: thr_val = float(self.threshold.get().strip())
            except ValueError:
                messagebox.showerror("Invalid Threshold",
                                     "Threshold অবশ্যই একটি সংখ্যা হতে হবে।")
                return

        offices_dir = os.path.join(out_folder, OUTPUT_SUBFOLDER)
        self.processing = True
        self.run_btn.configure(state="disabled", bg="#6C757D")
        self.progress.start(12)
        self._set_status("প্রক্রিয়াকরণ চলছে…")
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0","end")
        self.log_text.configure(state="disabled")

        def worker():
            try:
                master, n_customers, months, usage_cols, has_mob = process_files(
                    self.files, self._log)

                os.makedirs(offices_dir, exist_ok=True)
                self._log(f"\n🏢 অফিসওয়ারী আউটপুট ফোল্ডার:\n   {offices_dir}")

                results = []
                for code, name in OFFICES.items():
                    subset = subset_by_office(master, code)
                    if subset.empty:
                        self._log(f"   ⚠️ {name} ({code}): কোনো গ্রাহক পাওয়া যায়নি, এড়িয়ে যাওয়া হলো", "warn")
                        results.append((code, name, 0, 0))
                        continue
                    wb_o, mci_o = build_formatted_workbook(
                        subset, months, usage_cols, has_mob, sheet_title=name)
                    kept = len(subset)
                    if apply_flt:
                        _t, kept, _r = filter_low_usage(
                            wb_o, mci_o, thr_val, self._log, label=name)
                    office_path = os.path.join(offices_dir, f"{name}.xlsx")
                    wb_o.save(office_path)
                    self._log(f"   📥 {name}: total={len(subset)}, kept={kept}", "ok")
                    results.append((code, name, len(subset), kept))

                self._log(f"\n🎉 সম্পন্ন! মোট ইউনিক গ্রাহক: {n_customers}", "ok")
                made = sum(1 for r in results if r[2] > 0)
                self._set_status(f"✅ {made} office file(s) created", SUCCESS)

                lines = ["✅ ১০টি অফিস ফাইল তৈরি হয়েছে!\n",
                         f"মোট ইউনিক গ্রাহক: {n_customers}",
                         f"মাস: {', '.join(months)}"]
                if apply_flt:
                    lines.append(f"ফিল্টার থ্রেশহোল্ড: ≤ {thr_val}")
                lines.append(f"\nফোল্ডার:\n  {offices_dir}\n")
                for code, name, total, kept in results:
                    if total == 0:
                        lines.append(f"  • {name:<11} — (নেই)")
                    elif apply_flt:
                        lines.append(f"  • {name:<11} — total {total}, kept {kept}")
                    else:
                        lines.append(f"  • {name:<11} — {total} consumers")
                summary = "\n".join(lines)
                self.after(0, lambda: messagebox.showinfo("সম্পন্ন", summary))

                try:
                    if sys.platform == "win32":  os.startfile(offices_dir)
                    elif sys.platform == "darwin": subprocess.Popen(["open", offices_dir])
                    else: subprocess.Popen(["xdg-open", offices_dir])
                except: pass
            except Exception as e:
                tb = traceback.format_exc()
                self._log(f"❌ Error: {e}", "err")
                self._log(tb, "err")
                self._set_status("❌ Failed", ERROR)
                self.after(0, lambda: messagebox.showerror("ত্রুটি", str(e)))
            finally:
                self.processing = False
                self.after(0, lambda: (
                    self.run_btn.configure(state="normal", bg=DARK_BLUE),
                    self.progress.stop()))

        threading.Thread(target=worker, daemon=True).start()


# ================================================================
if __name__ == "__main__":
    app = App()
    app.mainloop()
